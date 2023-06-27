from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter

import CreateEvent
import Events
from Events import event

latestSchedule = [event]
month31 = [1, 3, 5, 7, 8, 10, 12]
month30 = [4, 6, 9, 11]

DATE_LOC = 4
NIGHT_SHIFT = "23:45"
MORNING_SHIFT = "06:15"
Day_SHIFT = "12:00"
EVENING_SHIFT = "18:00"
disallowed_chars_in_date = "\'\nאבגדהוש"


def removeCharsFromDate(date):
    for character in disallowed_chars_in_date:
        date = date.replace(character, "")
    return date


def getHeader(startT):
    if startT == NIGHT_SHIFT:
        return "Night shift"
    elif startT == MORNING_SHIFT:
        return "Morning shift"
    elif startT == Day_SHIFT:
        return "Day shift"
    else:
        return "Evening shift"


def analyzeExcel(fileName,user):
    try:
        wb = load_workbook(fileName)
        ws = wb.active
    except:
        print("Download was taking too long, check your connection")
    for col in range(1, 19):
        for row in range(4, 70):
            char = get_column_letter(col)
            if ws[char + str(row)].value == user:
                date = ws[char + str(DATE_LOC)].value
                date = removeCharsFromDate(date)

                # separate days from month and year and cast to int
                startDay = int(date.split('/')[0])
                startMonth = int(date.split('/')[1])
                year = int(date.split('/')[2])

                # separate hours from minutes and cast to int
                endT = ws[char + str(row + 1)].value
                endHour = int(endT.split(':')[0])
                endMinutes = int(endT.split(':')[1])

                startT = ws[get_column_letter(col + 1) + str(row + 1)].value
                startHour = int(startT.split(':')[0])
                startMinutes = int(startT.split(':')[1])

                # adjust dates in case of night shift
                if startHour > 21:
                    if startDay != 1:
                        startDay = startDay - 1
                    elif startMonth - 1 in month31:
                        startDay = 31
                        startMonth = startMonth - 1
                    elif startMonth - 1 in month30:
                        startDay = 30
                        startMonth = startMonth - 1
                    else:
                        startDay = 28
                        startMonth = 2

                    if startMonth in month31:
                        if startDay == 31:
                            endDay = 1
                            endMonth = startMonth + 1
                        else:
                            endDay = startDay + 1
                            endMonth = startMonth
                    elif startMonth in month30:
                        if startDay == 30:
                            endDay = 1
                            endMonth = startMonth + 1
                        else:
                            endDay = startDay + 1
                            endMonth = startMonth
                    elif startDay == 28:
                        endDay = 1
                        endMonth = 3
                else:
                    endDay = startDay
                    endMonth = startMonth

                summary = getHeader(startT)
                latestSchedule.append(Events.event(
                    startHour, startMinutes, endHour, endMinutes,
                    startDay, endDay, startMonth, endMonth, year, summary))
                print("Event was added")

    CreateEvent.insertEvents(latestSchedule)

